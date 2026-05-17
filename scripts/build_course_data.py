import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook
from geocode_shrines import slugify


WORKBOOK_PATH = Path("korean_catholic_holy_sites.xlsx")
OUTPUT_PATH = Path("src/data/courses.ts")
COURSE_SHEET_NAME = "추천코스"
SHRINE_SHEET_NAME = "성지목록"


def normalize_text(value: str) -> str:
    value = value or ""
    value = value.lower()
    value = re.sub(r"\([^)]*\)", "", value)
    value = re.sub(r"[^0-9a-z가-힣]+", "", value)
    return value


def ts_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def normalize_id(raw_id: str, name: str, used_ids: set[str]) -> str:
    base = raw_id or slugify(name)
    candidate = base
    index = 2
    while candidate in used_ids:
        candidate = f"{base}-{index}"
        index += 1
    used_ids.add(candidate)
    return candidate


def read_shrine_mappings(workbook):
    if SHRINE_SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"{SHRINE_SHEET_NAME} 시트를 찾지 못했습니다.")

    worksheet = workbook[SHRINE_SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"{SHRINE_SHEET_NAME} 시트가 비어 있습니다.")

    header = rows[0]
    index = {name: position for position, name in enumerate(header) if name is not None}
    required = ["성지명"]
    missing_columns = [name for name in required if name not in index]
    if missing_columns:
        raise SystemExit(f"{SHRINE_SHEET_NAME} 시트에 필수 컬럼이 없습니다: {', '.join(missing_columns)}")

    ids = set()
    exact_names = {}
    normalized_ids = {}
    normalized_names = {}
    used_ids = set()

    for row in rows[1:]:
        name = str(row[index["성지명"]] or "").strip()
        if not name:
            continue

        shrine_id = normalize_id("", name, used_ids)
        ids.add(shrine_id)
        exact_names[name] = shrine_id
        normalized_ids.setdefault(normalize_text(shrine_id), []).append(shrine_id)
        normalized_names.setdefault(normalize_text(name), []).append(shrine_id)

    return ids, exact_names, normalized_ids, normalized_names


def resolve_shrine_id(raw_id: str, raw_name: str, ids, exact_names, normalized_ids, normalized_names):
    raw_id = (raw_id or "").strip()
    raw_name = (raw_name or "").strip()

    if raw_id in ids:
        return raw_id

    normalized_id_candidates = normalized_ids.get(normalize_text(raw_id), [])
    if len(normalized_id_candidates) == 1:
        return normalized_id_candidates[0]

    if raw_name in exact_names:
        return exact_names[raw_name]

    candidates = normalized_names.get(normalize_text(raw_name), [])
    if len(candidates) == 1:
        return candidates[0]

    raise ValueError(f"성지 ID를 찾지 못했습니다: id={raw_id!r}, name={raw_name!r}")


def read_courses():
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"{WORKBOOK_PATH} 파일이 없습니다.")

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    if COURSE_SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"{COURSE_SHEET_NAME} 시트를 찾지 못했습니다.")

    worksheet = workbook[COURSE_SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    index = {name: position for position, name in enumerate(header) if name is not None}
    required = ["코스ID", "코스명", "지역", "테마", "권장일정", "이동수단", "코스설명", "순서", "성지ID", "성지명", "사용여부"]
    missing_columns = [name for name in required if name not in index]
    if missing_columns:
        raise SystemExit(f"추천코스 시트에 필수 컬럼이 없습니다: {', '.join(missing_columns)}")

    shrine_ids, exact_names, normalized_ids, normalized_names = read_shrine_mappings(workbook)
    grouped = OrderedDict()

    for row in rows[1:]:
        if not row:
            continue

        course_id = row[index["코스ID"]]
        if not course_id:
            continue

        enabled = str(row[index["사용여부"]] or "").strip().upper()
        if enabled != "Y":
            continue

        entry = grouped.setdefault(
            str(course_id).strip(),
            {
                "id": str(course_id).strip(),
                "title": str(row[index["코스명"]] or "").strip(),
                "region": str(row[index["지역"]] or "").strip(),
                "theme": str(row[index["테마"]] or "").strip(),
                "duration": str(row[index["권장일정"]] or "").strip(),
                "transport": str(row[index["이동수단"]] or "").strip(),
                "description": str(row[index["코스설명"]] or "").strip(),
                "stops": [],
            },
        )

        resolved_shrine_id = resolve_shrine_id(
            str(row[index["성지ID"]] or "").strip(),
            str(row[index["성지명"]] or "").strip(),
            shrine_ids,
            exact_names,
            normalized_ids,
            normalized_names,
        )

        entry["stops"].append((int(row[index["순서"]]), resolved_shrine_id))

    courses = []
    for entry in grouped.values():
        shrine_ids = [shrine_id for _, shrine_id in sorted(entry["stops"], key=lambda item: item[0])]
        courses.append(
            {
                "id": entry["id"],
                "title": entry["title"],
                "region": entry["region"],
                "theme": entry["theme"],
                "duration": entry["duration"],
                "transport": entry["transport"],
                "description": entry["description"],
                "shrineIds": shrine_ids,
            }
        )

    return courses


def write_ts(courses):
    lines = [
        "export type RecommendedCourse = {",
        "  id: string;",
        "  title: string;",
        "  region: string;",
        "  theme: string;",
        "  duration: string;",
        "  transport: string;",
        "  description: string;",
        "  shrineIds: string[];",
        "};",
        "",
        "export const recommendedCourses: RecommendedCourse[] = [",
    ]

    for course in courses:
        lines.extend(
            [
                "  {",
                f'    id: "{ts_string(course["id"])}",',
                f'    title: "{ts_string(course["title"])}",',
                f'    region: "{ts_string(course["region"])}",',
                f'    theme: "{ts_string(course["theme"])}",',
                f'    duration: "{ts_string(course["duration"])}",',
                f'    transport: "{ts_string(course["transport"])}",',
                f'    description: "{ts_string(course["description"])}",',
                "    shrineIds: [",
            ]
        )
        for shrine_id in course["shrineIds"]:
            lines.append(f'      "{ts_string(shrine_id)}",')
        lines.extend(["    ]", "  },"])

    if courses:
        lines[-1] = "  }"

    lines.extend(["];", ""])
    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main():
    courses = read_courses()
    write_ts(courses)
    print(f"{len(courses)}개 추천 코스를 {OUTPUT_PATH}에 반영했습니다.")


if __name__ == "__main__":
    raise SystemExit(main())
